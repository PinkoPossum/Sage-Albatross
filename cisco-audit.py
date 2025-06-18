import openpyxl
from netmiko import ConnectHandler
from getpass import getpass
from datetime import datetime
from openpyxl.styles import Font, PatternFill

def get_device_info():
    """
    Reads device IPs or domain names from an Excel file.
    """
    try:
        workbook = openpyxl.load_workbook('devices.xlsx')
        sheet = workbook.active
        # Assumes header is in row 1, devices start from row 2 in column A
        devices = [cell.value for cell in sheet['A'][1:] if cell.value]
        return devices
    except FileNotFoundError:
        print("Error: devices.xlsx not found. Please create it with device IPs/hostnames in column A.")
        return []

def connect_and_fetch(device, username, password):
    """
    Connects to a device via SSH and executes commands.
    """
    cisco_device = {
        'device_type': 'cisco_ios',
        'host': device,
        'username': username,
        'password': password,
    }

    commands = [
        'show version',
        'show interfaces status',
        'show interfaces switchport',
        'show interfaces trunk',
        'show ip route'
    ]
    output = {}

    try:
        with ConnectHandler(**cisco_device) as net_connect:
            for command in commands:
                output[command] = net_connect.send_command(command, read_timeout=20)
        return output
    except Exception as e:
        return f"Error connecting to {device}: {e}"

def main():
    """
    Main function to orchestrate the process.
    """
    devices = get_device_info()
    if not devices:
        return

    username = input("Enter your SSH username: ")
    password = getpass("Enter your SSH password: ")

    # Get the current time for the data timestamp within the file
    data_timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    output_workbook = openpyxl.Workbook()
    output_sheet = output_workbook.active
    output_sheet.title = "Device Command Outputs"
    output_sheet.append(["Device", "Timestamp", "Command", "Output"])

    red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

    for device in devices:
        print(f"\nConnecting to {device}...")
        command_outputs = connect_and_fetch(device, username, password)

        if isinstance(command_outputs, str):
            print(command_outputs)
            output_sheet.append([device, data_timestamp, "Connection Status", "Needs manual verification"])
            continue

        print(f"Successfully connected to {device}. Fetching command outputs...")
        # A single timestamp is used for all commands for a single device run
        for command, output in command_outputs.items():
            # Split the output by lines to insert it into multiple rows if needed
            output_lines = output.strip().splitlines()
            
            # Write the header row for the command
            output_sheet.append([device, data_timestamp, command, output_lines[0] if output_lines else ""])
            
            # Write the rest of the output
            if len(output_lines) > 1:
                for line in output_lines[1:]:
                    output_sheet.append(["", "", "", line])

            # Highlight down interfaces
            if command == 'show interfaces status':
                # Find the starting row for the current command's output
                start_row = output_sheet.max_row - len(output_lines) + 1
                for i in range(len(output_lines)):
                    # Check for 'down' or 'disabled' in the interface status line
                    if ' down ' in output_lines[i] or 'disabled' in output_lines[i]:
                        # Highlight the entire row for that interface's data
                        for col in range(1, 5):
                            output_sheet.cell(row=start_row + i, column=col).fill = red_fill

    # --- FILENAME FORMATTING CHANGED HERE ---
    # Format the date for the filename as dd-mm-yy
    file_date_format = datetime.now().strftime("%d-%m-%y")
    output_filename = f"{file_date_format}-audit.xlsx"
    
    output_workbook.save(output_filename)
    print(f"\nScript finished. Output saved to {output_filename}")

if __name__ == "__main__":
    main()
